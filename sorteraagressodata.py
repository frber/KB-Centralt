import openpyxl
import math
import datetime
import os
from openpyxl.styles import Font


class SorteraAgressodata:

    def __init__(self, wb4):
        self.wb4 = wb4
        self.sheet = self.wb4["Klistra in Agressodata här"]


    def nollstall(self):
        for sheets in self.wb4.worksheets:
            if sheets == self.wb4["Klistra in Agressodata här"]:
                continue
            else:
                sheets.delete_rows(5, 15000)


    def sortera_data(self):

        rad = 4
        for sheet in self.wb4.worksheets:
            #print(sheet, sheet.max_row)
            enhet = sheet['A3'].value
            kontoklass = sheet['B3'].value

            for row in self.sheet['A1:A10000']:
                for cell in row:
                    if cell.value != None:
                        konto = cell.value
                        kontot = cell.offset(column=1).value
                        resenhet = cell.offset(column=2).value
                        projekt = cell.offset(column=3).value
                        projektt = cell.offset(column=4).value
                        belopp = cell.offset(column=5).value
                        dnr = cell.offset(column=6).value
                        sluttdat = cell.offset(column=7).value
                        projtyp = cell.offset(column=8).value
                        #print(str(konto)[:1])
                        if enhet == resenhet and str(konto)[:1] == str(kontoklass):
                            if kontoklass == 1 and konto >= 1618 and konto <= 1679:
                                sheet.cell(row=sheet.max_row + 1, column=1).value = konto
                                sheet.cell(row=sheet.max_row, column=2).value = kontot
                                sheet.cell(row=sheet.max_row, column=3).value = resenhet
                                sheet.cell(row=sheet.max_row, column=4).value = projekt
                                sheet.cell(row=sheet.max_row, column=5).value = projektt
                                sheet.cell(row=sheet.max_row, column=6).value = belopp
                                sheet.cell(row=sheet.max_row, column=7).value = dnr
                                sheet.cell(row=sheet.max_row, column=8).value = sluttdat
                                sheet.cell(row=sheet.max_row, column=9).value = projtyp
                            if kontoklass == 2 and konto >= 2711 and konto <= 2777:
                                sheet.cell(row=sheet.max_row + 1, column=1).value = konto
                                sheet.cell(row=sheet.max_row, column=2).value = kontot
                                sheet.cell(row=sheet.max_row, column=3).value = resenhet
                                sheet.cell(row=sheet.max_row, column=4).value = projekt
                                sheet.cell(row=sheet.max_row, column=5).value = projektt
                                sheet.cell(row=sheet.max_row, column=6).value = belopp
                                sheet.cell(row=sheet.max_row, column=7).value = dnr
                                sheet.cell(row=sheet.max_row, column=8).value = sluttdat
                                sheet.cell(row=sheet.max_row, column=9).value = projtyp



    def spara(self):
        try:
            self.wb.save(self.agressodata)
            self.wb.close()
        except PermissionError:
            return True
        #os.startfile(self.agressodata)


















