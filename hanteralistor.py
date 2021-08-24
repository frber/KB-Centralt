import openpyxl
from outputfil import OutputFil

class HanteraListor:

    #wb2 = openpyxl.load_workbook(OutputFil.outputfil, data_only=True)
    #ws_ejberper = OutputFil.ws_ejberper
    #ws_trasiga = OutputFil.ws_trasiga
    #ws_forstora = OutputFil.ws_forstora

    #outputfil = "Docs\\Berperdata.xlsx"
    #wb2 = openpyxl.load_workbook(outputfil, data_only=True)
    #ws_ejberper = wb2["Filer som ej bedöms vara berper"]
    #ws_trasiga = wb2['Excelfiler som ej kunde öppnas']
    #ws_forstora = wb2['För stora excelfiler']


    #rad = 1

    def __init__(self, lista_alla_filer, lista_berpers, lista_trasiga_filer, lista_stora_excelfiler, wb2):
        self.lista_alla_filer = lista_alla_filer
        self.lista_berpers = lista_berpers
        self.lista_trasiga_filer = lista_trasiga_filer
        self.list_stora_excelfiler = lista_stora_excelfiler
        self.wb2 = wb2
        self.ws_ejberper = self.wb2["Filer som ej bedöms vara berper"]
        self.ws_trasiga = self.wb2['Excelfiler som ej kunde öppnas']
        self.ws_forstora = self.wb2['För stora excelfiler']


    def kontrollera_listor(self):
        # Filer som ej bedöms vara berpers med sortering av filformat.
        diff_alla_filer = set(self.lista_alla_filer) - set(self.lista_berpers)
        diff_alla_filer = list(diff_alla_filer)
        diff_alla_filer.sort(reverse=True, key=lambda x: x[-4:])

        # Förhindrar att en fil som är trasig dyker upp i listan för filer som ej bedöms vara berper
        for x in diff_alla_filer:
            if x not in self.lista_trasiga_filer:
                self.ws_ejberper.cell(row=self.ws_ejberper.max_row + 1, column=1).value = x

        for y in self.lista_trasiga_filer:
            #self.rad+=1
            self.ws_trasiga.cell(row=self.ws_trasiga.max_row  + 1, column=1).value = y

        for z in self.list_stora_excelfiler:
            self.ws_forstora.cell(row=self.ws_forstora.max_row + 1, column=1).value = z

        self.wb2.close()








