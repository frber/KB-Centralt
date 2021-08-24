import os
from collections import Counter
import time
from tkinter import *
import tkinter as tk
import tkinter.ttk as ttk
from tkinter import Tk, Label, Button
from tkinter import filedialog
from tkinter import messagebox
import threading
from PIL import ImageTk, Image
import shutil
import openpyxl

#Egenskapade klasser.
from hanterafil import HanteraFil
from kontrollera import Kontrollera
from outputfil import OutputFil
from hanteraoutput import HanteraOutput
from hanteralistor import HanteraListor
from outputmodel import OutputModel
from slaihop import SlaIhop
from sorteraagressodata import SorteraAgressodata

class Granssnitt:

    def __init__(self, master, tab1, tab2):
        self.tab1 = tab1
        self.tab2 = tab2
        self.master = master
        master.title("KB 1.8")
        master.configure(bg="gray22")


        self.folderpath = tk.StringVar()
        self.folderpath2 = tk.StringVar()

        bakgrund = "white smoke"

        #TAB1---------------------------------------------------------------------------------------
        # Bild.
        self.bild = ImageTk.PhotoImage(Image.open("Docs\\ltu3.jpg"))
        self.label_bild = Label(tab1, image=self.bild, bg=bakgrund)
        # Label EoP.
        self.label_eop = Label(tab1, text="Kontroll av berper", font="Arial 13", bg=bakgrund)
        self.label_eop.place(y=20, x=340, anchor="center")
        # Label "Kontroll av berper".
        #self.label_eop = Label(tab1, text="Kontroll av berper", font="Arial 13", bg=bakgrund)
        #self.label_eop.place(y=45, x=340, anchor="center")

        # Knapp, klistra in data från Agresso.
        self.knapp_agressodata = Button(tab1, text="Klistra in data från Agresso", command=self.oppna_agressodata)
        self.knapp_agressodata.place(y=100, x=240, anchor="center")
        self.knapp_agressodata.config(height=1, width=20)
        # Knapp sortera Agressodata
        self.knapp_sortera = Button(tab1, text="Sortera Agressodata", command=self.thread2)
        self.knapp_sortera.place(y=140, x=240, anchor="center")
        self.knapp_sortera.config(height=1, width=20)


        #Knapp, label till välj mapp.
        self.knapp_valj = Button(tab1, text="Välj mapp för e-bokslut", command=self.get_folder_path)
        self.knapp_valj.place(y=100, x=420, anchor="center")
        self.knapp_valj.config(height=1, width=20)
        self.mapp_label = Label(tab1, text="Ingen filväg vald", font="Arial 7", bg=bakgrund)
        self.mapp_label.place(y=265, x=340, anchor="center")
        # Knapp till huvudprogram.
        self.knapp_starta = Button(tab1, text="Starta", command=self.thread)
        self.knapp_starta.place(y=140, x=420, anchor="center")
        self.knapp_starta.config(height=1, width=20)

        # Label antal kontrollerade berpers.
        self.label_raknade_berpers = Label(tab1, text="Kontrollerade berpes: 0", font="Arial 8", bg=bakgrund)
        self.label_raknade_berpers.place(y=18, x=600, anchor="center")
        # Label avvikande berpers.
        #self.label_avvikande_berpers = Label(tab1, text="Antal avvikelser: 0", font="Arial 8", bg=bakgrund)
        #self.label_avvikande_berpers.place(y=35, x=613, anchor="center")

        # Label hittad fil.
        self.label_fil = Label(tab1, text="", font="Arial 15", bg=bakgrund)
        self.label_fil.place(y=210, x=340, anchor="center")
        #Progressbar.
        self.s = ttk.Style()
        self.s.theme_use("clam")
        self.s.configure("blue.Horizontal.TProgressbar", foreground='navy', background='navy')
        self.prog_bar = ttk.Progressbar(tab1, style="blue.Horizontal.TProgressbar", orient=HORIZONTAL, length=600, maximum=100, mode='determinate')
        self.prog_bar.place(y=240, x=350, anchor="center")
        # Label kontakt.
        self.label_kontakt = Label(tab1, text="Kontakt: fredrik.bergstrom@ltu.se, anknytning 3982", font="Arial 8",bg=bakgrund)
        self.label_kontakt.place(y=315, x=125, anchor="center")

        # TAB2---------------------------------------------------------------------------------------
        # Bild.
        self.label_bild.place(y=35, x=40, anchor="center")
        self.label_bild = Label(tab2, image=self.bild, bg=bakgrund)
        self.label_bild.place(y=35, x=40, anchor="center")

        # Knapp nollställ Agressodata
        self.knapp_nollstall_agresso = Button(tab2, text="Nollställ dokument Agressodata", command=self.nollstall_agressodata)
        self.knapp_nollstall_agresso.place(y=100, x=120, anchor="center")
        self.knapp_nollstall_agresso.config(height=1, width=30)

        # Knapp nollställ Berperdata
        self.knapp_nollstall_berperdata = Button(tab2, text="Nollställ dokument för Berperdata",command=self.nollstall_berperdata_knapp)
        self.knapp_nollstall_berperdata.place(y=140, x=120, anchor="center")
        self.knapp_nollstall_berperdata.config(height=1, width=30)


        #self.mapp_label2 = Label(tab2, text="Ingen fil vald", font="Arial 7", bg=bakgrund)
        #self.mapp_label2.place(y=220, x=340, anchor="center")
        # Knapp och command till För över data.
        #self.knapp_ihop = Button(tab2, text="TEST", command=self.test)
        #self.knapp_ihop.place(y=140, x=65, anchor="center")
        #self.knapp_ihop.config(height=1, width=15)
        # Label kontakt.
        self.label_kontakt = Label(tab2, text="Kontakt: fredrik.bergstrom@ltu.se, anknytning 3982", font="Arial 8",bg=bakgrund)
        self.label_kontakt.place(y=315, x=125, anchor="center")


    def get_folder_path(self):
        #Frågar, hämtar och anger vald filväg.
        folder_selected = filedialog.askdirectory()
        self.folderpath.set(folder_selected)
        self.mapp_label["text"] = folder_selected

    def get_file_path(self):
        # Frågar, hämtar och anger vald filväg.
        file_selected = filedialog.askopenfilename()
        self.folderpath2.set(file_selected)
        self.mapp_label2["text"] = file_selected

    def thread(self):
        #Använder en annan thread så att gränssnittet inte fryser medans huvudprogrammet körs.
        #Startar lokalt i en egen metod eftersom detta måste instansieras på nytt, annars: RuntimeError: threads can only be started once.
        t = threading.Thread(target=self.huvudprog, daemon=True)
        t.start()

    def thread2(self):
        # Använder en annan thread så att gränssnittet inte fryser medans huvudprogrammet körs.
        # Startar lokalt i en egen metod eftersom detta måste instansieras på nytt, annars: RuntimeError: threads can only be started once.
        t = threading.Thread(target=self.sortera, daemon=True)
        t.start()


    def oppna_agressodata(self):
        os.startfile('Docs\\Agressodata.xlsx')

    def nollstall_berperdata(self):
        try:
            os.remove('Docs\\Berperdata.xlsx')
            shutil.copy('Docs\\Orginal\\Berperdata.xlsx', 'Docs\\Berperdata.xlsx')
        except PermissionError:
            return True

    def nollstall_agressodata(self):
        try:
            os.remove('Docs\\Agressodata.xlsx')
            shutil.copy('Docs\\Orginal\\Agressodata.xlsx', 'Docs\\Agressodata.xlsx')
            messagebox.showinfo("Dokument nollställt", "Dokumentet för Agressodata är nu nollställt!")
        except PermissionError:
            messagebox.showerror("OBS!", "Du har filen för Agressodata öppen, stäng ned den och börja om.")

    def nollstall_berperdata_knapp(self):
        try:
            os.remove('Docs\\Berperdata.xlsx')
            shutil.copy('Docs\\Orginal\\Berperdata.xlsx', 'Docs\\Berperdata.xlsx')
            messagebox.showinfo("Dokument nollställt", "Dokumentet för Berperdata är nu nollställt!")
        except PermissionError:
            messagebox.showerror("OBS!", "Du har filen för Berperdata öppen, stäng ned den och börja om.")


    def sortera(self):
        agressodata = 'Docs\\Agressodata.xlsx'
        wb4 = openpyxl.load_workbook(agressodata, data_only=True)
        agressodatasort = SorteraAgressodata(wb4)
        agressodatasort.nollstall()
        agressodatasort.sortera_data()
        try:
            wb4.save(agressodata)
            wb4.close()
            messagebox.showinfo("Agressodata sorterat", "Dokumentet för Agressodata är nu sorterat!")
        except PermissionError:
            messagebox.showerror("OBS!", "Du har filen för Agressodata öppen, stäng ned den och börja om.")


 
    def huvudprog(self):
        lista_alla_filer = []
        lista_berpers = []
        lista_trasiga_filer = []
        lista_stora_excelfiler = []
        
        rakna_berper = 0
        rakna_avvikande_berper = 0
        #outputfil = OutputFil()
        if self.nollstall_berperdata():
             messagebox.showerror("OBS!", "Du har filen för Berperdata öppen, stäng ned den och börja om.")
        else:
        #outputfil.nollstall()
        #if outputfil.nollstall():
            #messagebox.showerror("OBS!", "Du har filen för Berperdata öppen, stäng ned den innan du fortsätter")
        #else
            #outputfil.nollstall()
            sok_filvag = self.folderpath.get()
            berperdata = 'Docs\\Berperdata.xlsx'
            wb2 = openpyxl.load_workbook(berperdata, data_only=True)


            for root, dirs, files in os.walk(sok_filvag):
                for fil in files:
                    hantera_fil = HanteraFil(root, fil)
                    filplats = hantera_fil.filplats
                    lista_alla_filer.append(filplats)
                    if filplats.split(".")[-1] == "xlsx" and not hantera_fil.storlek():
                        lista_stora_excelfiler.append(filplats)
                    sheet = hantera_fil.avgor_om_berper()
                    if sheet == "fel":
                        lista_trasiga_filer.append(filplats)
                    else:
                        if sheet != None:
                            self.prog_bar.start(1)
                            lista_berpers.append(filplats)
                            filnamn_clean = hantera_fil.filnamn_clean()
                            self.label_fil["text"] = filnamn_clean
                            berper = Kontrollera(sheet, filnamn_clean, filplats)
                            berper = berper.validera()
                            output = HanteraOutput(berper, wb2, berperdata)
                            output.skriv_output()
                            rakna_berper += 1
                            self.label_raknade_berpers["text"] = "Kontrollerade berpers: "+str(rakna_berper)
                            #if output.avvikade_berper():
                                #rakna_avvikande_berper +=1
                                #self.label_avvikande_berpers["text"] = "Antal avvikelser: " + str(rakna_avvikande_berper)

            listor = HanteraListor(lista_alla_filer, lista_berpers, lista_trasiga_filer, lista_stora_excelfiler, wb2)
            listor.kontrollera_listor()
            try:
                wb2.save(berperdata)
                wb2.close()
                self.prog_bar.stop()
                if rakna_berper > 0:
                    self.label_fil["text"] = "Klar"
                    output.starta()
                    self.ihopslag()
            except PermissionError:
                messagebox.showerror("OBS!", "Du har filen för Berperdata öppen, stäng ned den och börja om.")
            #output.spara_stang()
            #output.kontroll_dokument()


    def ihopslag(self):
        #filvag_slaihop = "Ihopslagen browser.xlsx" #self.folderpath2.get()
        agressodata = 'Docs\\Agressodata.xlsx'
        wb3 = openpyxl.load_workbook(agressodata, data_only=True)
        ihop = SlaIhop(wb3)
        #ihop.clear()
        #ihop.skapa_rubriker()
        ihop.kontroll_browserdata()
        ihop.con()
        ihop.slutkontroll()
        ihop.lagg_in_lankar()
        try:
            wb3.save(agressodata)
            wb3.close()
            os.startfile(agressodata)
        except PermissionError:
            messagebox.showerror("OBS!", "Du har filen för Agressodata öppen, stäng ned den och börja om.")

        #if ihop.spara():
            #messagebox.showerror("OBS!", "Du har filen för Agressodata öppen, stäng ned den och börja om.")
        #else:
            #ihop.spara()
            #ihop.oppna()




def main():
    root = Tk()
    root.geometry("680x350")
    root.resizable(width=False, height=False)
    root.iconbitmap('Docs\\favicon.ico')


    tabcontrol = ttk.Notebook(root)

    tab1 = Frame(tabcontrol, bg="white smoke")
    tabcontrol.add(tab1, text="Kontroll")
    tabcontrol.pack(expand=1, fill="both")
    ttk.Style().configure("TNotebook", bg="black")

    tab2 = Frame(tabcontrol, bg="white smoke")
    tabcontrol.add(tab2, text="Hantera dokument")


    app = Granssnitt(root, tab1, tab2)


    root.mainloop()

if __name__ == "__main__":
    main()


# ATT GÖRA:
# Vänd - till + för belopp periodisering i avvikerlser2.
# Tid istället för storlek.
# Slå ihop med browserfråga.
# Gör loggarna för filerna i separata flikar.
# Skapa ny outputfil istället för att använda befintlig.
# Kontrollera rätt tecken - borde kunna göras direkt i browserfrågan eftersom både konto och belopp med tecken finns med.
# Kontrollera slutdatum
# Färdig länk till platina beroende på DNR?
# Se över filer som inte kunde öppnas
# Säkrare att bygga in int+kst = 0 och per > 0 i browserfråga än hämta från berper?





# Hur ger man över mappen och dokumentet till rev?
# Flytta om och kopiera till en komplett mapp för revisorer?
# Standardisera browserfrågan - effektiverisera? Köra alla ebok samtidigt?






